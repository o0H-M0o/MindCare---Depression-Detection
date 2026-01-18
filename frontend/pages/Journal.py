"""
Journal page - User input for daily reflections
"""

import streamlit as st
from datetime import datetime
import requests
import sys
import re
import pandas as pd
from io import BytesIO
from pathlib import Path
import openpyxl
import time

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))

from utils.db_client import DBClient
from utils.auth import init_auth_service
from utils.auth_sidebar import render_auth_sidebar

# Add backend path for imports
backend_path = Path(__file__).parent.parent.parent / "backend"
sys.path.append(str(backend_path))

from model.preprocess import validate_and_clean_entry

# Import backend modules for immediate BDI analysis
from model.llm_model import BDIAssessmentModel
from model.preprocess import clean_entry
from utils.score_bdi import calculate_total_score, get_depression_category
from model.sentiment_model import SentimentAnalyzer

# Template URL
TEMPLATE_URL = "https://amnfjolprnnqfqskrlrh.supabase.co/storage/v1/object/public/excel_template/Template.xlsx"

def limit_records_by_word_count(records, text_key: str, max_words: int) -> list:
    """Limit records to the most recent items whose combined word count stays within max_words."""
    limited = []
    total_words = 0
    for record in records:
        text = str(record.get(text_key, "")).strip()
        if not text:
            continue
        words = len(text.split())
        if total_words + words > max_words:
            break
        limited.append(record)
        total_words += words
    return limited

def parse_whatsapp_chat(content: str, user_name: str = None) -> pd.DataFrame:
    """Parse WhatsApp chat export and return DataFrame with date, time, text columns."""
    lines = content.strip().split('\n')
    parsed_messages = []
    
    # Regex pattern for WhatsApp format: [DD/MM/YYYY, HH:MM:SS] Name: Message
    pattern = r'^\[(\d{1,2}/\d{1,2}/\d{4}), (\d{1,2}:\d{1,2}:\d{1,2})\] (.+?):\s*(.+)$'
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        match = re.match(pattern, line)
        if match:
            date_str, time_str, sender, message = match.groups()
            
            # Filter by user name if provided (case-insensitive)
            if user_name and sender.strip().lower() != user_name:
                continue
                
            # Convert date format from DD/MM/YYYY to YYYY-MM-DD
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                formatted_date = date_obj.strftime('%Y-%m-%d')
            except ValueError:
                formatted_date = date_str  # Keep original if parsing fails
                
            # Preprocess the message
            try:
                cleaned_message = clean_entry(message)
                if not cleaned_message:
                    cleaned_message = message  # Fallback to original if cleaning fails
            except Exception:
                cleaned_message = message  # Fallback to original
                
            parsed_messages.append({
                'Date': formatted_date,
                'Time': time_str,
                'Text': cleaned_message
            })
    
    return pd.DataFrame(parsed_messages)

def fill_excel_template(messages_df: pd.DataFrame) -> BytesIO:
    """Fill the Excel template with parsed messages and return as BytesIO."""
    try:
        # Download template
        response = requests.get(TEMPLATE_URL)
        response.raise_for_status()
        
        # Load template into memory
        template_buffer = BytesIO(response.content)
        
        # Load the workbook
        workbook = openpyxl.load_workbook(template_buffer)
        sheet = workbook.active
        
        # Clear existing data (starting from row 2, assuming row 1 has headers)
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
        
        # Write our data starting from row 2 (ignore DataFrame index to avoid blank rows)
        write_df = messages_df[["Date", "Time", "Text"]].reset_index(drop=True)
        for excel_row, msg in enumerate(write_df.itertuples(index=False), start=2):
            sheet.cell(row=excel_row, column=1).value = msg.Date
            sheet.cell(row=excel_row, column=2).value = msg.Time
            sheet.cell(row=excel_row, column=3).value = msg.Text

        # Trim any remaining empty rows below the written data
        data_end_row = 1 + len(write_df)  # header row + data rows
        current_max_row = sheet.max_row
        if current_max_row > data_end_row:
            sheet.delete_rows(data_end_row + 1, current_max_row - data_end_row)
        
        # Save to output buffer
        output_buffer = BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer
        
    except Exception as e:
        st.error(f"Error filling Excel template: {str(e)}")
        return None

# Page config
st.set_page_config(
    page_title="My Analyser",
    page_icon="📝",
    layout="wide"
)

# Initialize auth and render auth-aware sidebar
auth_service = init_auth_service()
render_auth_sidebar(auth_service)

# Enforce only individual users can access Journal
auth_service.require_role(['individual'])

# Initialize client with authenticated user id
current_user = auth_service.get_current_user()
if not current_user:
    st.error("Authentication required")
    st.stop()

user_id = current_user['id']
db_client = DBClient(user_id=user_id)

# Load analysis models once per session for better performance
if 'bdi_model' not in st.session_state:
    st.session_state.bdi_model = BDIAssessmentModel()

if 'sentiment_analyzer' not in st.session_state:
    st.session_state.sentiment_analyzer = SentimentAnalyzer()

st.title("📝 My Analyser")
st.write("Share your thoughts and feelings. Your entries help track your mental well-being over time.")

tab1, tab2 = st.tabs(["Upload for Analysis", "Quick Entry"])

with tab1:
    st.subheader("Upload for Analysis")
    st.write("You can add writing you’d like to reflect on by uploading a file. This could be chat histories, social media posts, or notes you’ve written.")
    st.write("The system will return the overall analysis for the uploaded file.")
    with st.expander("📄 Option 1: WhatsApp Chat History"):
        st.write("**Optional: Choose Messages to Include**")
        st.write("If the chat includes multiple people, you can choose which person's messages you'd like to include for reflection.")
        user_name = st.text_input("Your username in the chat history (optional):", key="whatsapp_user_name")
        st.divider()
        with st.expander("💡 How to get my WhatsApp chat history?"):
            st.write("To export a WhatsApp chat, open the chat, tap the contact/group name (iPhone) or More options (Android), select Export Chat, choose Without Media, then share the resulting .txt file via email, cloud storage, or messaging apps")
            image_path = Path(__file__).parent.parent / "utils" / "image" / "export_chat.png"
            st.image(str(image_path), caption="Steps to export WhatsApp chat history")
        st.caption("The system will process recent messages up to about 800 words.")
        uploaded_file = st.file_uploader("Choose a WhatsApp export file (.txt)", type=['txt'], key="whatsapp_file_uploader")

        if uploaded_file is not None:
            if st.button("🕵️ Save & Analyse", key="analyse_whatsapp_messages", width='stretch'):
                with st.spinner("🔍 Analyzing text... Please stay on the page while analysis completes."):
                    try:
                        # Step 1: Read and decode file
                        try:
                            file_content = uploaded_file.read().decode("utf-8")
                        except UnicodeDecodeError:
                            st.error("❌ Error: Unable to read the file. Please ensure it's a valid text file encoded in UTF-8.")
                            st.info("💡 Tip: WhatsApp exports are usually UTF-8 encoded. Try saving the file with UTF-8 encoding.")
                            st.stop()
                        except Exception as e:
                            st.error(f"❌ Error reading file: {str(e)}")
                            st.stop()
                        
                        # Step 2: Parse the chat
                        try:
                            messages_df = parse_whatsapp_chat(file_content, user_name.lower().strip() if user_name and user_name.strip() else None)
                            
                            # Reverse the order so latest messages appear first
                            messages_df = messages_df[::-1]

                            # Check if we have valid data
                            if messages_df.empty or 'Date' not in messages_df.columns or 'Time' not in messages_df.columns or 'Text' not in messages_df.columns:
                                st.warning("⚠️ No messages found in the uploaded file.")
                                st.info("💡 This could happen if:\n- The file format doesn't match WhatsApp export\n- No messages from the specified user (if name was provided)\n- The chat export is empty")
                                st.stop()

                            # Limit by total word count (most recent first)
                            recent_records = messages_df.to_dict(orient='records')
                            total_words_all = sum(len(str(r.get("Text", "")).strip().split()) for r in recent_records)
                            limited_records = limit_records_by_word_count(recent_records, "Text", max_words=800)
                            if total_words_all > 800:
                                st.warning(f"This system only analyses the most recent rows up to 800 words. Rows beyond this limit will be excluded.")
                            if not limited_records:
                                st.warning("⚠️ No messages were kept within the 800-word limit.")
                                st.stop()
                            messages_df = pd.DataFrame(limited_records)
                            
                            # Sort ascending by Date and Time
                            messages_df = messages_df.sort_values(by=['Date', 'Time'], ascending=True)
                            
                        except Exception as e:
                            st.error(f"❌ Error parsing WhatsApp chat: {str(e)}")
                            st.info("💡 Please ensure your file is in the correct WhatsApp export format: '[DD/MM/YYYY, HH:MM:SS] Name: Message'")
                            st.stop()
                        
                        # Step 3: Initialize models
                        try:
                            bdi_model = st.session_state.bdi_model
                            sentiment_analyzer = st.session_state.sentiment_analyzer
                        except Exception as e:
                            st.error(f"❌ Error initializing analysis models: {str(e)}")
                            st.info("💡 This might be a temporary issue. Please try again later.")
                            st.stop()

                        # Step 3b: Require minimum words in combined text before saving
                        joined_preview = "\n".join([str(t).strip() for t in messages_df["Text"].tolist() if str(t).strip()])
                        cleaned_preview = clean_entry(joined_preview) or joined_preview
                        preview_word_count = len(cleaned_preview.split())
                        if preview_word_count < 50:
                            st.warning("⚠️ The uploaded file doesn't contain enough text. Please upload at least 50 words before saving.")
                            st.stop()

                        # Step 4: Save and analyze messages
                        analysis_delay_seconds = 1.2
                        total = len(messages_df)
                        progress = st.progress(0)
                        status = st.empty()

                        # Step 4a: Save all messages as file_pending first
                        status.write("💾 Saving all messages...")
                        entries_to_analyze = []
                        uploaded_file_name = f"{uploaded_file.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        used_datetimes: set[tuple[str, str]] = set()

                        for i, msg in enumerate(messages_df.reset_index(drop=True).to_dict(orient='records'), start=1):
                            date_str = str(msg.get('Date', '')).strip()
                            time_str = str(msg.get('Time', '')).strip()
                            entry_text = str(msg.get('Text', '')).strip()

                            if not entry_text:
                                continue

                            try:
                                entry_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                            except Exception:
                                entry_dt = datetime.now()

                            # Avoid (user_id, entry_date, entry_time) collisions
                            unique_date = entry_dt.date().isoformat()
                            unique_time = entry_dt.time().isoformat()
                            while (unique_date, unique_time) in used_datetimes:
                                entry_dt = entry_dt.replace(microsecond=(entry_dt.microsecond + 1) % 1_000_000)
                                unique_date = entry_dt.date().isoformat()
                                unique_time = entry_dt.time().isoformat()
                            used_datetimes.add((unique_date, unique_time))

                            try:
                                entry_id = db_client.save_journal_entry(
                                    text=entry_text,
                                    entry_date=entry_dt,
                                    entry_type='by_upload',
                                    uploaded_file=uploaded_file_name,
                                )
                            except Exception as e:
                                error_str = str(e)
                                if 'duplicate key value violates unique constraint' in error_str and 'journal_entries_user_datetime_unique' in error_str:
                                    # Skip duplicate entries silently
                                    continue
                                else:
                                    st.error(f"❌ Error saving message to database: {error_str}")
                                    st.info("💡 Please check your database connection and try again.")
                                    st.stop()

                            if entry_id:
                                db_client.update_journal_entry_status(entry_id, 'file_pending')
                                entries_to_analyze.append((entry_id, entry_text))

                            progress.progress((i / max(total, 1)) * 0.5)

                        # Step 4b: Analyze combined messages once (chunked if too long)

                        if not entries_to_analyze:
                            st.warning("⚠️ No messages were saved for analysis.")
                            st.stop()

                        # Join all saved messages for a single analysis pass
                        joined_text = "\n".join([t for _, t in entries_to_analyze if t])
                        cleaned_joined = clean_entry(joined_text) or joined_text

                        # Chunk by preprocessing limit (max words)
                        max_words = 400
                        words = cleaned_joined.split()
                        chunks = []
                        for start in range(0, len(words), max_words):
                            chunk = " ".join(words[start:start + max_words]).strip()
                            if chunk and len(chunk.split()) >= 50:
                                chunks.append(chunk)
                        
                        # Limit to maximum 2 chunks to prevent excessive processing
                        chunks = chunks[:2]

                        if not chunks:
                            st.warning("⚠️ Combined text is empty after preprocessing.")
                            st.stop()

                        assessment_results_list = []
                        pos_scores = []
                        neu_scores = []
                        neg_scores = []

                        total_steps = max(total + len(chunks), 1)
                        for idx, chunk in enumerate(chunks, start=1):
                            try:
                                assessment_results = bdi_model.assess_all_symptoms([chunk])
                                assessment_results_list.append(assessment_results)

                                sentiment_result = sentiment_analyzer.analyze(chunk)
                                if sentiment_result:
                                    scores = sentiment_result.get('scores', {})
                                    pos_scores.append(float(scores.get('Positive', 0.0)))
                                    neu_scores.append(float(scores.get('Neutral', 0.0)))
                                    neg_scores.append(float(scores.get('Negative', 0.0)))
                            except Exception:
                                # Continue even if a chunk fails
                                pass

                            progress.progress((total + idx) / total_steps)
                            time.sleep(analysis_delay_seconds)

                        # Aggregate assessment results across chunks
                        if assessment_results_list:
                            qids = list(assessment_results_list[0].keys())
                            aggregated_results = {}
                            for qid in qids:
                                levels = [r.get(qid, {}).get('level', 0) for r in assessment_results_list]
                                avg_level = round(sum(levels) / max(len(levels), 1))
                                base = assessment_results_list[0].get(qid, {})
                                aggregated_results[qid] = {
                                    "level": avg_level,
                                    "reason": f"Aggregated from {len(assessment_results_list)} part(s)",
                                    "symptom": base.get("symptom"),
                                }

                            total_score = calculate_total_score(aggregated_results)
                            avg_category = get_depression_category(total_score)
                        else:
                            aggregated_results = {}
                            total_score = 0
                            avg_category = 'N/A'

                        # Average sentiment across chunks
                        if pos_scores:
                            avg_pos = sum(pos_scores) / len(pos_scores)
                            avg_neu = sum(neu_scores) / len(neu_scores)
                            avg_neg = sum(neg_scores) / len(neg_scores)

                            if avg_pos > avg_neu and avg_pos > avg_neg:
                                avg_sentiment = 'Positive'
                            elif avg_neu > avg_pos and avg_neu > avg_neg:
                                avg_sentiment = 'Neutral'
                            else:
                                avg_sentiment = 'Negative'
                        else:
                            avg_pos = avg_neu = avg_neg = 0.0
                            avg_sentiment = 'N/A'

                        # Save the same results to all file_pending entries for this upload
                        for entry_id, _ in entries_to_analyze:
                            try:
                                if aggregated_results:
                                    db_client.save_assessment(
                                        entry_id=entry_id,
                                        assessment_data=aggregated_results,
                                        total_score=total_score,
                                        category=avg_category,
                                    )

                                if avg_sentiment != 'N/A':
                                    db_client.save_sentiment_analysis(
                                        entry_id=entry_id,
                                        top_label=avg_sentiment,
                                        positive_score=avg_pos,
                                        neutral_score=avg_neu,
                                        negative_score=avg_neg,
                                    )

                                db_client.update_journal_entry_status(entry_id, 'completed')
                            except Exception as e:
                                db_client.update_journal_entry_status(entry_id, 'failed', str(e))

                        status.write("✅ Finished saving and analyzing uploaded messages.")

                        # Conclusion like Tab 2
                        sentiment_messages = {
                            "Positive": "<strong>Positive</strong> mood 😊",
                            "Neutral": "<strong>Neutral</strong> mood 😐",
                            "Negative": "<strong>Low</strong> mood 😔"
                        }
                        sentiment_text = sentiment_messages.get(avg_sentiment, "Mood analysis unavailable 💭")
                        
                        category_messages = {
                            "Minimal": "<strong>No signs</strong> of depressive symptoms",
                            "Mild": "<strong>Mild</strong> depressive symptoms detected",
                            "Moderate": "<strong>Moderate</strong> depressive symptoms detected",
                            "Severe": "<strong>Severe</strong> depressive symptoms detected"
                        }
                        category_message = category_messages.get(avg_category, avg_category)
                        
                        result_text = f"""{category_message}
{sentiment_text}"""
                        
                        st.write("**Analysis Summary:**")
                        st.write("")
                        # Color-coded conclusion box
                        if avg_category == "Minimal":
                            st.markdown(f'<div style="background-color: #f0f9f0; color: #2d5a2d; padding: 10px; border-radius: 5px; border-left: 5px solid #4caf50;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Mild":
                            st.markdown(f'<div style="background-color: #fefde7; color: #bf360c; padding: 10px; border-radius: 5px; border-left: 5px solid #ffb74d;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Moderate":
                            st.markdown(f'<div style="background-color: #fff4e6; color: #c62828; padding: 10px; border-radius: 5px; border-left: 5px solid #ff7043;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Severe":
                            st.markdown(f'<div style="background-color: #fef2f2; color: #ad1457; padding: 10px; border-radius: 5px; border-left: 5px solid #e91e63;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        else:
                            st.markdown(f'<div style="background-color: #f0f8ff; color: #1565c0; padding: 10px; border-radius: 5px; border-left: 5px solid #42a5f5;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        st.write("")      
                    except Exception as e:
                        st.error(f"❌ Unexpected error during processing: {str(e)}")
                        st.info("💡 Please try again or contact support if the problem persists.")
                        st.write("")
    
    with st.expander("📝 Option 2: Other Sources"):
        st.write("Use this option if your writing comes from other places, such as social media posts, notes, or conversations from other apps")
        
        st.write("- Please add your text using the provided **Excel template**")
        st.write("- Use one row for each message or thought.")
        st.write("")
        
        st.write("**Example**")
        example_data = {
            'Date': ['2025-01-02', '2025-01-03', '2025-01-06'],
            'Time': ['22:30', '11:00', '19:00'],
            'Text': ['So excited about my upcoming vacation!', 'Had the best coffee date with my best friend today', 'The sunset by the beach was breathtaking']
        }
        st.table(example_data)
        
        template_url="https://amnfjolprnnqfqskrlrh.supabase.co/storage/v1/object/public/excel_template/Template.xlsx"
        response = requests.get(template_url)
        template_bytes = response.content
        st.download_button(
            label="📥 Download Excel Template",
            data=template_bytes,
            width='stretch',
            file_name="template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        uploaded_excel = st.file_uploader("📤 Upload Filled Template", type=['xlsx', 'xls'], key="excel_uploader")
        if uploaded_excel is not None:

            if st.button("🕵️ Save & Analyse", key="analyse_excel_template", width='stretch'):
                with st.spinner("🔍 Analyzing text... Please stay on the page while analysis completes."):
                    try:
                        # Step 1: Read Excel
                        try:
                            df = pd.read_excel(uploaded_excel)
                        except Exception as e:
                            st.error(f"❌ Error reading the Excel file: {str(e)}")
                            st.info("💡 Please upload a valid .xlsx/.xls file based on the template.")
                            st.stop()

                        if df is None or df.empty:
                            st.error("❌ The uploaded Excel file is empty.")
                            st.info("💡 Please fill at least one row with Date, Time, and Text.")
                            st.stop()

                        # Step 2: Validate required columns (case-insensitive)
                        df.columns = [str(c).strip() for c in df.columns]
                        col_map = {str(c).strip().lower(): str(c).strip() for c in df.columns}
                        required = ['date', 'time', 'text']
                        missing = [c for c in required if c not in col_map]

                        if missing:
                            st.error("❌ Invalid template format.")
                            st.info(
                                "💡 Your Excel must contain these columns (exact names): Date, Time, Text.\n"
                                f"Missing: {', '.join(missing)}"
                            )
                            st.stop()

                        df = df.rename(columns={
                            col_map['date']: 'Date',
                            col_map['time']: 'Time',
                            col_map['text']: 'Text',
                        })

                        # Step 3: Normalize rows and enforce (Date, Time, Text)
                        upload_ts = datetime.now()
                        records = []
                        errors = 0
                        for _, row in df.iterrows():
                            date_val = row.get('Date')
                            time_val = row.get('Time')
                            text_val = row.get('Text')

                            if pd.isna(text_val) or not str(text_val).strip():
                                continue

                            # Date (auto-fill if missing)
                            try:
                                if pd.isna(date_val):
                                    date_iso = upload_ts.date().isoformat()
                                elif hasattr(date_val, 'date'):
                                    date_iso = date_val.date().isoformat()
                                else:
                                    date_iso = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date().isoformat()
                            except Exception:
                                errors += 1
                                continue

                            # Time (auto-fill if missing)
                            try:
                                if pd.isna(time_val):
                                    time_str = upload_ts.strftime('%H:%M:%S')
                                elif hasattr(time_val, 'strftime'):
                                    time_str = time_val.strftime('%H:%M:%S')
                                elif isinstance(time_val, (int, float)) and 0 <= float(time_val) < 1:
                                    total_seconds = int(round(float(time_val) * 86400))
                                    hh = total_seconds // 3600
                                    mm = (total_seconds % 3600) // 60
                                    ss = total_seconds % 60
                                    time_str = f"{hh:02d}:{mm:02d}:{ss:02d}"
                                else:
                                    time_str = str(time_val).strip()
                                    # Accept HH:MM and normalize to HH:MM:SS
                                    if re.match(r'^\d{1,2}:\d{2}$', time_str):
                                        time_str = f"{time_str}:00"

                                # Validate final time format
                                if not re.match(r'^\d{1,2}:\d{2}:\d{2}$', time_str):
                                    raise ValueError("Invalid Time")
                            except Exception:
                                errors += 1
                                continue

                            # Text (cleaned)
                            try:
                                cleaned_text = clean_entry(str(text_val).strip())
                                if not cleaned_text:
                                    cleaned_text = str(text_val).strip()
                            except Exception:
                                cleaned_text = str(text_val).strip()

                            records.append({'Date': date_iso, 'Time': time_str, 'Text': cleaned_text})

                        if not records:
                            st.error("❌ No valid rows found in the Excel file.")
                            st.info("💡 Ensure each row has valid Date (YYYY-MM-DD), Time (HH:MM or HH:MM:SS), and Text.")
                            st.stop()

                        if errors:
                            st.warning(f"⚠️ Skipped {errors} invalid row(s).")

                        # Limit by total word count (most recent by Date/Time)
                        records = sorted(records, key=lambda r: (r['Date'], r['Time']), reverse=True)
                        total_words_all = sum(len(str(r.get("Text", "")).strip().split()) for r in records)
                        records = limit_records_by_word_count(records, "Text", max_words=800)
                        if total_words_all > 800:
                            st.warning(f"This system only analyses the most recent rows up to 800 words. Rows beyond this limit will be excluded.")
                        records = sorted(records, key=lambda r: (r['Date'], r['Time']), reverse=False)

                        # Step 4: Initialize models
                        try:
                            bdi_model = st.session_state.bdi_model
                            sentiment_analyzer = st.session_state.sentiment_analyzer
                        except Exception as e:
                            st.error(f"❌ Error initializing analysis models: {str(e)}")
                            st.info("💡 This might be a temporary issue. Please try again later.")
                            st.stop()

                        # Step 4b: Require minimum words in combined text before saving
                        joined_preview = "\n".join([str(r.get('Text', '')).strip() for r in records if str(r.get('Text', '')).strip()])
                        cleaned_preview = clean_entry(joined_preview) or joined_preview
                        preview_word_count = len(cleaned_preview.split())
                        if preview_word_count < 50:
                            st.warning("⚠️ The uploaded file doesn't contain enough text. Please upload at least 50 words before saving.")
                            st.stop()

                        # Step 5: Save and analyze
                        analysis_delay_seconds = 1.2
                        total = len(records)
                        progress = st.progress(0)
                        status = st.empty()

                        status.write("💾 Saving all rows...")
                        entries_to_analyze = []
                        uploaded_file_name = f"{uploaded_excel.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        used_datetimes: set[tuple[str, str]] = set()

                        for i, msg in enumerate(records, start=1):
                            date_str = str(msg.get('Date', '')).strip()
                            time_str = str(msg.get('Time', '')).strip()
                            entry_text = str(msg.get('Text', '')).strip()

                            if not entry_text:
                                continue

                            try:
                                entry_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                            except Exception:
                                entry_dt = datetime.now()

                            # Avoid (user_id, entry_date, entry_time) collisions
                            unique_date = entry_dt.date().isoformat()
                            unique_time = entry_dt.time().isoformat()
                            while (unique_date, unique_time) in used_datetimes:
                                entry_dt = entry_dt.replace(microsecond=(entry_dt.microsecond + 1) % 1_000_000)
                                unique_date = entry_dt.date().isoformat()
                                unique_time = entry_dt.time().isoformat()
                            used_datetimes.add((unique_date, unique_time))

                            try:
                                entry_id = db_client.save_journal_entry(
                                    text=entry_text,
                                    entry_date=entry_dt,
                                    entry_type='by_upload',
                                    uploaded_file=uploaded_file_name,
                                )
                            except Exception as e:
                                st.error(f"❌ Error saving row to database: {str(e)}")
                                st.info("💡 Please check your database connection and try again.")
                                st.stop()

                            if entry_id:
                                db_client.update_journal_entry_status(entry_id, 'file_pending')
                                entries_to_analyze.append((entry_id, entry_text))

                            progress.progress((i / max(total, 1)) * 0.5)

                        if not entries_to_analyze:
                            st.warning("⚠️ No rows were saved for analysis.")
                            st.stop()

                        joined_text = "\n".join([t for _, t in entries_to_analyze if t])
                        cleaned_joined = clean_entry(joined_text) or joined_text

                        max_words = 400
                        words = cleaned_joined.split()
                        chunks = []
                        for start in range(0, len(words), max_words):
                            chunk = " ".join(words[start:start + max_words]).strip()
                            if chunk and len(chunk.split()) >= 50:
                                chunks.append(chunk)
                        
                        # Limit to maximum 2 chunks to prevent excessive processing
                        chunks = chunks[:2]

                        if not chunks:
                            st.warning("⚠️ Combined text is empty after preprocessing.")
                            st.stop()

                        assessment_results_list = []
                        pos_scores = []
                        neu_scores = []
                        neg_scores = []

                        total_steps = max(total + len(chunks), 1)
                        for idx, chunk in enumerate(chunks, start=1):
                            try:
                                assessment_results = bdi_model.assess_all_symptoms([chunk])
                                assessment_results_list.append(assessment_results)

                                sentiment_result = sentiment_analyzer.analyze(chunk)
                                if sentiment_result:
                                    scores = sentiment_result.get('scores', {})
                                    pos_scores.append(float(scores.get('Positive', 0.0)))
                                    neu_scores.append(float(scores.get('Neutral', 0.0)))
                                    neg_scores.append(float(scores.get('Negative', 0.0)))
                            except Exception:
                                pass

                            progress.progress((total + idx) / total_steps)
                            time.sleep(analysis_delay_seconds)

                        if assessment_results_list:
                            qids = list(assessment_results_list[0].keys())
                            aggregated_results = {}
                            for qid in qids:
                                levels = [r.get(qid, {}).get('level', 0) for r in assessment_results_list]
                                avg_level = round(sum(levels) / max(len(levels), 1))
                                base = assessment_results_list[0].get(qid, {})
                                aggregated_results[qid] = {
                                    "level": avg_level,
                                    "reason": f"Aggregated from {len(assessment_results_list)} part(s)",
                                    "symptom": base.get("symptom"),
                                }

                            total_score = calculate_total_score(aggregated_results)
                            avg_category = get_depression_category(total_score)
                        else:
                            aggregated_results = {}
                            total_score = 0
                            avg_category = 'N/A'

                        if pos_scores:
                            avg_pos = sum(pos_scores) / len(pos_scores)
                            avg_neu = sum(neu_scores) / len(neu_scores)
                            avg_neg = sum(neg_scores) / len(neg_scores)

                            if avg_pos > avg_neu and avg_pos > avg_neg:
                                avg_sentiment = 'Positive'
                            elif avg_neu > avg_pos and avg_neu > avg_neg:
                                avg_sentiment = 'Neutral'
                            else:
                                avg_sentiment = 'Negative'
                        else:
                            avg_pos = avg_neu = avg_neg = 0.0
                            avg_sentiment = 'N/A'

                        for entry_id, _ in entries_to_analyze:
                            try:
                                if aggregated_results:
                                    db_client.save_assessment(
                                        entry_id=entry_id,
                                        assessment_data=aggregated_results,
                                        total_score=total_score,
                                        category=avg_category,
                                    )

                                if avg_sentiment != 'N/A':
                                    db_client.save_sentiment_analysis(
                                        entry_id=entry_id,
                                        top_label=avg_sentiment,
                                        positive_score=avg_pos,
                                        neutral_score=avg_neu,
                                        negative_score=avg_neg,
                                    )

                                db_client.update_journal_entry_status(entry_id, 'completed')
                            except Exception as e:
                                db_client.update_journal_entry_status(entry_id, 'failed', str(e))

                        status.write("✅ Finished saving and analyzing uploaded Excel rows.")

                        sentiment_messages = {
                            "Positive": "<strong>Positive</strong> mood 😊",
                            "Neutral": "<strong>Neutral</strong> mood 😐",
                            "Negative": "<strong>Low</strong> mood 😔"
                        }
                        sentiment_text = sentiment_messages.get(avg_sentiment, "Mood analysis unavailable 💭")

                        category_messages = {
                            "Minimal": "<strong>No signs</strong> of depressive symptoms",
                            "Mild": "<strong>Mild</strong> depressive symptoms detected",
                            "Moderate": "<strong>Moderate</strong> depressive symptoms detected",
                            "Severe": "<strong>Severe</strong> depressive symptoms detected"
                        }
                        category_message = category_messages.get(avg_category, avg_category)
                        result_text = f"""{category_message}
{sentiment_text}"""

                        st.write("**Analysis Summary:**")
                        st.write("")
                        if avg_category == "Minimal":
                            st.markdown(f'<div style="background-color: #f0f9f0; color: #2d5a2d; padding: 10px; border-radius: 5px; border-left: 5px solid #4caf50;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Mild":
                            st.markdown(f'<div style="background-color: #fefde7; color: #bf360c; padding: 10px; border-radius: 5px; border-left: 5px solid #ffb74d;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Moderate":
                            st.markdown(f'<div style="background-color: #fff4e6; color: #c62828; padding: 10px; border-radius: 5px; border-left: 5px solid #ff7043;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        elif avg_category == "Severe":
                            st.markdown(f'<div style="background-color: #fef2f2; color: #ad1457; padding: 10px; border-radius: 5px; border-left: 5px solid #e91e63;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        else:
                            st.markdown(f'<div style="background-color: #f0f8ff; color: #1565c0; padding: 10px; border-radius: 5px; border-left: 5px solid #42a5f5;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                        st.write("")
                    except Exception as e:
                        st.error(f"❌ Unexpected error during processing: {str(e)}")
                        st.info("💡 Please try again or contact support if the problem persists.")
                        st.write("")

    st.divider()
    st.subheader("📊 Recent Analyses")

    recent_uploads = db_client.get_recent_entries(days=None, limit=100, entry_type='by_upload')  # Get more to group properly

    if recent_uploads:
        # Group by uploaded_file
        from collections import defaultdict
        grouped = defaultdict(list)
        for entry in recent_uploads:
            file_key = entry.get('uploaded_file', 'Unknown')
            grouped[file_key].append(entry)

        for file_name, entries in grouped.items():
            # Extract original file name and timestamp
            if '_' in file_name and len(file_name) > 16:
                file_base = file_name[:-16]
                timestamp = file_name[-15:]
                try:
                    year = timestamp[:4]
                    month = timestamp[4:6]
                    day = timestamp[6:8]
                    hour = timestamp[9:11]
                    minute = timestamp[11:13]
                    second = timestamp[13:]
                    formatted_datetime = f"{year}-{month}-{day} {hour}:{minute}:{second}"
                    display_name = f"{file_base} (uploaded by {formatted_datetime})"
                except:
                    display_name = file_name
            else:
                display_name = file_name
            with st.expander(f"📁 {display_name}"):
                # Collect data for averages and table
                table_data = []
                total_scores = []
                pos_scores = []
                neu_scores = []
                neg_scores = []

                for entry in entries:
                    assessment = db_client.get_assessment_by_entry(entry['id'])
                    sentiment = db_client.get_sentiment_by_entry(entry['id'])
                    
                    dep_cat = assessment['category'] if assessment else 'N/A'
                    sent_label = sentiment['top_label'] if sentiment else 'N/A'
                    
                    table_data.append({
                        'Message': entry['text'][:50] + '...' if len(entry['text']) > 50 else entry['text']
                    })
                    
                    if assessment:
                        total_scores.append(assessment['total_score'])
                    if sentiment:
                        pos_scores.append(sentiment['positive_score'])
                        neu_scores.append(sentiment['neutral_score'])
                        neg_scores.append(sentiment['negative_score'])

                # Average results
                if total_scores:
                    avg_score = sum(total_scores) / len(total_scores)
                    avg_category = get_depression_category(avg_score)
                else:
                    avg_category = 'N/A'

                if pos_scores:
                    avg_pos = sum(pos_scores) / len(pos_scores)
                    avg_neu = sum(neu_scores) / len(neu_scores)
                    avg_neg = sum(neg_scores) / len(neg_scores)

                    if avg_pos > avg_neu and avg_pos > avg_neg:
                        avg_sentiment = 'Positive'
                    elif avg_neu > avg_pos and avg_neu > avg_neg:
                        avg_sentiment = 'Neutral'
                    else:
                        avg_sentiment = 'Negative'
                else:
                    avg_sentiment = 'N/A'

                # Conclusion like Tab 2
                sentiment_messages = {
                    "Positive": "<strong>Positive</strong> mood 😊",
                    "Neutral": "<strong>Neutral</strong> mood 😐",
                    "Negative": "<strong>Low</strong> mood 😔"
                }
                sentiment_text = sentiment_messages.get(avg_sentiment, "Mood analysis unavailable 💭")
                
                category_messages = {
                    "Minimal": "<strong>No signs</strong> of depressive symptoms",
                    "Mild": "<strong>Mild</strong> depressive symptoms detected",
                    "Moderate": "<strong>Moderate</strong> depressive symptoms detected",
                    "Severe": "<strong>Severe</strong> depressive symptoms detected"
                }
                category_message = category_messages.get(avg_category, avg_category)
                
                result_text = f"""{category_message}
{sentiment_text}"""
                
                # Color-coded conclusion box
                if avg_category == "Minimal":
                    st.markdown(f'<div style="background-color: #f0f9f0; color: #2d5a2d; padding: 10px; border-radius: 5px; border-left: 5px solid #4caf50;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                elif avg_category == "Mild":
                    st.markdown(f'<div style="background-color: #fefde7; color: #bf360c; padding: 10px; border-radius: 5px; border-left: 5px solid #ffb74d;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                elif avg_category == "Moderate":
                    st.markdown(f'<div style="background-color: #fff4e6; color: #c62828; padding: 10px; border-radius: 5px; border-left: 5px solid #ff7043;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                elif avg_category == "Severe":
                    st.markdown(f'<div style="background-color: #fef2f2; color: #ad1457; padding: 10px; border-radius: 5px; border-left: 5px solid #e91e63;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="background-color: #f0f8ff; color: #1565c0; padding: 10px; border-radius: 5px; border-left: 5px solid #42a5f5;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)

                # Table: Message
                st.write("")
                st.dataframe(table_data, width='stretch', hide_index=True)
    else:
        st.info("No recent analyses. Upload and analyze a file to see results here!")

with tab2:
    # Main journal input
    st.subheader("How are you feeling today?")

    journal_text = st.text_area(
        "💡 Tip: Aim for 200-400 words for meaningful insights",
        height=200,
        placeholder="Write about your day, your thoughts, feelings, or anything on your mind...",
        help="Your entries are private and will be analyzed to help understand your mood patterns.",
        key="journal_input"
    )

    # Real-time word counter
    if journal_text:
        word_count = len(journal_text.split())
        max_words = 400
        
        if word_count < 50:
            st.write(f"⚠️ **{word_count} / 50 words** - Please write at least 50 words for meaningful analysis")
        elif word_count > max_words:
            st.write(f"⚠️ **{word_count} / {max_words} words** - Please keep your entry under {max_words} words")
        elif word_count > max_words * 0.9:  # Warning at 90%
            st.write(f"📝 **{word_count} / {max_words} words** - Approaching word limit")
        else:
            st.write(f"📝 **{word_count} / {max_words} words**")
    else:
        pass

    # Create placeholder immediately below text box for loading and results
    result_placeholder = st.empty()

    # Clear analysis results if text has changed
    if 'last_analyzed_text' in st.session_state and 'latest_analysis_result' in st.session_state:
        if st.session_state['last_analyzed_text'] != journal_text:
            # Text has changed, clear the results
            del st.session_state['latest_analysis_result']
            del st.session_state['last_analyzed_text']

    # Display stored analysis results if they exist
    if 'latest_analysis_result' in st.session_state and st.session_state['latest_analysis_result']:
        result_data = st.session_state['latest_analysis_result']
        
        # Display the stored result
        category = result_data.get('category')
        sentiment_label = result_data.get('sentiment_label')
        
        # Prepare sentiment text
        if sentiment_label:
            sentiment_messages = {
                "Positive": "<strong>Positive</strong> mood 😊",
                "Neutral": "<strong>Neutral</strong> mood 😐",
                "Negative": "<strong>Low</strong> mood 😔"
            }
            sentiment_text = sentiment_messages.get(sentiment_label, sentiment_label)
        else:
            sentiment_text = "Mood analysis unavailable 💭"
        
        # Format the result text
        category_messages = {
            "Minimal": "<strong>No signs</strong> of depressive symptoms",
            "Mild": "<strong>Mild</strong> depressive symptoms detected",
            "Moderate": "<strong>Moderate</strong> depressive symptoms detected",
            "Severe": "<strong>Severe</strong> depressive symptoms detected"
        }
        
        category_message = category_messages.get(category, category)
        
        result_text = f"""{category_message}
    {sentiment_text}"""
        
        # Display with color based on BDI category
        if category == "Minimal":
            st.markdown(f'<div style="background-color: #f0f9f0; color: #2d5a2d; padding: 10px; border-radius: 5px; border-left: 5px solid #4caf50;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
        elif category == "Mild":
            st.markdown(f'<div style="background-color: #fefde7; color: #bf360c; padding: 10px; border-radius: 5px; border-left: 5px solid #ffb74d;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
        elif category == "Moderate":
            st.markdown(f'<div style="background-color: #fff4e6; color: #c62828; padding: 10px; border-radius: 5px; border-left: 5px solid #ff7043;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
        elif category == "Severe":
            st.markdown(f'<div style="background-color: #fef2f2; color: #ad1457; padding: 10px; border-radius: 5px; border-left: 5px solid #e91e63;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="background-color: #f0f8ff; color: #1565c0; padding: 10px; border-radius: 5px; border-left: 5px solid #42a5f5;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
        
        st.write("")

    # Save button appears here
    save_button_container = st.container()

    # # Entry metadata
    # col1, col2 = st.columns(2)

    # with col1:
    #     entry_date = st.date_input(
    #         "Date",
    #         value=datetime.now(),
    #         help="Date of this journal entry"
    #     )

    # with col2:
    #     entry_time = st.time_input(
    #         "Time",
    #         value=datetime.now().time(),
    #         help="Time of this journal entry"
    #     )
    # # Entry metadata (generated fresh on each save to ensure unique IDs)
    # entry_date = datetime.now()
    # entry_time = datetime.now().time()
    # entry_datetime = datetime.combine(entry_date, entry_time).replace(microsecond=0)


    def validate_entry(text: str):
        """
        Validate journal entry via backend `validate_and_clean_entry`.

        Returns a `ValidationResult` (preferred) or the legacy (bool, str) tuple.
        """
        return validate_and_clean_entry(text)


    # Only one button now - Save button (placed in container after results area)
    with save_button_container:
        save_clicked = st.button("🕵️‍♂️ Save & Analyse", type="secondary", width='stretch')

    if save_clicked:
        # Generate completely fresh datetime for each save to ensure unique ID generation
        entry_datetime = datetime.combine(datetime.now(), datetime.now().time()).replace(microsecond=0)
        
        result_placeholder.empty()  # Clear previous results when saving again
        res = validate_entry(journal_text)

        # Handle old tuple-based result or new ValidationResult
        if isinstance(res, tuple):
            is_valid, result = res
            if not is_valid:
                st.error(result)
                # stop here
                pass
            else:
                cleaned_text = result
                proceed_save = True
        else:
            # assume ValidationResult
            if not getattr(res, 'success', False):
                st.error(getattr(res, 'message', 'Invalid entry'))
                proceed_save = False
                cleaned_text = None
            else:
                cleaned_text = res.cleaned_text
                proceed_save = True

        if not cleaned_text or not proceed_save:
            # nothing to do
            pass
        else:
            # Show spinning loading when user saves entry successfully
            with result_placeholder:
                with st.spinner("Saving your entry..."):
                    # Save the original user text to the DB
                    entry_id = db_client.save_journal_entry(
                        text=journal_text,
                        entry_date=entry_datetime,
                        entry_type='by_typing'
                    )

            if entry_id:
                # Show spinning loading for analysis
                # Initialize result variables
                assessment_saved = False
                category = None
                sentiment_result = None
                
                with result_placeholder:
                    with st.spinner("🔍 Analyzing text... Please stay on the page while analysis completes."):
                        try:
                            # Clean the text for analysis
                            cleaned_text = clean_entry(journal_text)
                            texts = [cleaned_text] if cleaned_text else [journal_text]
                            
                            # Initialize BDI model and assess
                            bdi_model = st.session_state.bdi_model
                            assessment_results = bdi_model.assess_all_symptoms(texts)
                            
                            # Calculate scores
                            total_score = calculate_total_score(assessment_results)
                            category = get_depression_category(total_score)
                            
                            # Save assessment to database
                            assessment_saved = db_client.save_assessment(
                                entry_id=entry_id,
                                assessment_data=assessment_results,
                                total_score=total_score,
                                category=category
                            )
                            
                            # Perform sentiment analysis
                            try:
                                sentiment_analyzer = st.session_state.sentiment_analyzer
                                sentiment_result = sentiment_analyzer.analyze(cleaned_text)
                                
                                if sentiment_result:
                                    # Save sentiment to database
                                    scores = sentiment_result.get('scores', {})
                                    sentiment_data = {
                                        'entry_id': entry_id,
                                        'user_id': db_client.user_id,
                                        'top_label': sentiment_result.get('label'),
                                        'positive_score': float(scores.get('Positive', 0.0)),
                                        'neutral_score': float(scores.get('Neutral', 0.0)),
                                        'negative_score': float(scores.get('Negative', 0.0))
                                    }
                                    db_client.supabase.table('sentiment_analysis').insert(sentiment_data).execute()
                                    print(f"✅ Sentiment saved: {sentiment_result.get('label')}")
                            except Exception as e:
                                print(f"⚠️ Sentiment analysis error: {e}")
                                sentiment_result = None
                            
                            # Update entry status to completed
                            db_client.supabase.table('journal_entries').update({
                                'analysis_status': 'completed'
                            }).eq('id', entry_id).execute()
                            
                        except Exception as e:
                            print(f"❌ Immediate analysis error: {e}")
                            assessment_saved = False
                            category = None
                            sentiment_result = None
                
                # Store the result in session state for persistent display
                if assessment_saved and category:
                    # Store results in session state
                    st.session_state['latest_analysis_result'] = {
                        'category': category,
                        'sentiment_label': sentiment_result.get('label') if sentiment_result else None
                    }
                    # Store the analyzed text to track changes
                    st.session_state['last_analyzed_text'] = journal_text
                    # Rerun to display the results
                    st.rerun()
                    # Note: Cannot clear input here due to Streamlit session state restrictions
                    # User can manually clear the input or we can implement a different approach
                else:
                    with result_placeholder:
                        st.warning("Entry saved but analysis couldn't be completed. Results will be available soon.")
            else:
                with result_placeholder:
                    st.error("Failed to save entry. Please try again.")

    # Display recent entries
    st.divider()
    st.subheader("📖 Recent Entries")

    recent_entries = db_client.get_recent_entries(days=None, limit=5, entry_type='by_typing')

    # Check for and process any pending entries from previous sessions (like interrupted WhatsApp uploads)
    pending_entries = db_client.get_recent_entries(days=None, limit=50, entry_type=None)  # Get more to check all types
    pending_to_process = [e for e in pending_entries if e.get('analysis_status') == 'pending']

    if pending_to_process:
        with st.spinner("🔄 Processing pending analyses from previous sessions..."):
            bdi_model = st.session_state.bdi_model
            sentiment_analyzer = st.session_state.sentiment_analyzer
            
            for entry in pending_to_process:
                try:
                    db_client.update_journal_entry_status(entry['id'], 'processing')
                    
                    cleaned_text = clean_entry(entry['text'])
                    texts = [cleaned_text] if cleaned_text else [entry['text']]
                    
                    assessment_results = bdi_model.assess_all_symptoms(texts)
                    total_score = calculate_total_score(assessment_results)
                    category = get_depression_category(total_score)
                    
                    db_client.save_assessment(
                        entry_id=entry['id'],
                        assessment_data=assessment_results,
                        total_score=total_score,
                        category=category,
                    )
                    
                    try:
                        sentiment_result = sentiment_analyzer.analyze(cleaned_text or entry['text'])
                        if sentiment_result:
                            scores = sentiment_result.get('scores', {})
                            db_client.save_sentiment_analysis(
                                entry_id=entry['id'],
                                top_label=sentiment_result.get('label'),
                                positive_score=float(scores.get('Positive', 0.0)),
                                neutral_score=float(scores.get('Neutral', 0.0)),
                                negative_score=float(scores.get('Negative', 0.0)),
                            )
                    except Exception:
                        pass
                    
                    db_client.update_journal_entry_status(entry['id'], 'completed')
                    
                except Exception as e:
                    db_client.update_journal_entry_status(entry['id'], 'failed', str(e))
        
        st.success(f"✅ Processed {len(pending_to_process)} pending analyses!")
        st.rerun()  # Refresh to show updated results

    if recent_entries:
        for entry in recent_entries:
            # Format date and time for display
            entry_time = entry.get('time', 'N/A')
            # Format time to show only HH:MM (remove seconds)
            if entry_time != 'N/A' and ':' in entry_time:
                time_parts = entry_time.split(':')
                entry_time = f"{time_parts[0]}:{time_parts[1]}"  # HH:MM
            
            # Simple title with date and time
            display_title = f"📅 {entry['date']}  \n⏰ {entry_time}"
            
            with st.expander(display_title):
                st.write(entry['text'])
                
                # Show BDI score if available
                assessment = db_client.get_assessment_by_entry(entry['id'])
                sentiment = db_client.get_sentiment_by_entry(entry['id'])
                
                if assessment:
                    category = assessment['category']
                    
                    # Prepare sentiment text
                    if sentiment:
                        sentiment_label = sentiment.get('top_label', 'Unknown')
                        sentiment_score = sentiment.get('positive_score', 0) if sentiment_label == 'Positive' else \
                                        sentiment.get('neutral_score', 0) if sentiment_label == 'Neutral' else \
                                        sentiment.get('negative_score', 0)
                        
                        sentiment_messages = {
                            "Positive": "<strong>Positive</strong> mood 😊",
                            "Neutral": "<strong>Neutral</strong> mood 😐",
                            "Negative": "<strong>Low</strong> mood 😔"
                        }
                        
                        sentiment_text = sentiment_messages.get(sentiment_label, sentiment_label)
                    else:
                        sentiment_text = "Mood analysis unavailable 💭"
                    
                    # Format the result text (same as main analysis)
                    category_messages = {
                        "Minimal": "<strong>No signs</strong> of depressive symptoms",
                        "Mild": "<strong>Mild</strong> depressive symptoms detected",
                        "Moderate": "<strong>Moderate</strong> depressive symptoms detected",
                        "Severe": "<strong>Severe</strong> depressive symptoms detected"
                    }
                    
                    category_message = category_messages.get(category, category)
                    
                    result_text = f"""{category_message}
    {sentiment_text}"""
                    
                    # Display with same color coding as main analysis
                    if category == "Minimal":
                        st.markdown(f'<div style="background-color: #f0f9f0; color: #2d5a2d; padding: 10px; border-radius: 5px; border-left: 5px solid #4caf50;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                    elif category == "Mild":
                        st.markdown(f'<div style="background-color: #fefde7; color: #bf360c; padding: 10px; border-radius: 5px; border-left: 5px solid #ffb74d;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                    elif category == "Moderate":
                        st.markdown(f'<div style="background-color: #fff4e6; color: #c62828; padding: 10px; border-radius: 5px; border-left: 5px solid #ff7043;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                    elif category == "Severe":
                        st.markdown(f'<div style="background-color: #fef2f2; color: #ad1457; padding: 10px; border-radius: 5px; border-left: 5px solid #e91e63;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div style="background-color: #f0f8ff; color: #1565c0; padding: 10px; border-radius: 5px; border-left: 5px solid #42a5f5;">{result_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
                    
                    st.write("")  
    else:
        st.info("No recent entries. Start writing to track your well-being!")


